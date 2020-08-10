import requests
import re
import time
from selenium import webdriver
from openpyxl import load_workbook
from lxml import etree
from queue import Queue
from threading import Thread
import threading
from time import sleep

class DaZhongDianPing():
    def __init__(self):
        #excel的行数
        self.flag = 2
        # 页面字体大小
        self.font_size = 14
        # 页面引用的 css 文件
        self.css = None
        # 商家地址使用的 svg 文件
        self.address_svg = None
        # 商家电话使用的 svg 文件
        self.tell_svg = None
        self.homeurl=None
        self.maxpage=None
        # 判断手机的正则
        self.phone_rule = '^(\\+\\d{2}-)?(\\d{2,3}-)?([1][3,4,5,7,8,9][0-9]\\d{8})$'
        # 生成一个已存在的workbook对象
        self.wb = load_workbook(".\\result.xlsx")
        # 激活表格
        self.wbw = self.wb.active

        self.referer=None
        self.timeout = 10
        self.headers = {
            'Connection': 'keep-alive',
            'Pragma': 'no-cache',
            'Cache-Control': 'no-cache',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.89 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Referer': self.referer,
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Cookie': ''
        }
    #通过webdriver登陆大众点评，获取登陆的cookie，否则无法获取完整电话
    def getCookies(self):
        print(self.headers['Cookie'])
        if self.headers['Cookie'] is not '':
            return
        brower = webdriver.Chrome('.\\chromedriver_win32\\chromedriver.exe')
        url = "http://www.dianping.com/"
        brower.get("https://account.dianping.com/login?redir=http%3A%2F%2Fwww.dianping.com%2F")
        cookie = ''
        while True:
            print("Please login in dianping.com!")
            sleep(3)
            if brower.current_url == url:
                tbCookies = brower.get_cookies()
                brower.quit()
                break
        for i in tbCookies:
            name = i['name']
            value = i['value']
            cookie = cookie + name + '=' + value + ';'
        self.headers['Cookie'] = cookie
        print(self.headers['Cookie'])
    # 从搜索页上获取每个店的的主页url
    def get_url(self):
        #从搜索页上获取每个店的的主页url
        for i in range(1,int(self.maxpage)+1):
            url = self.homeurl+'p'+str(i)
            print(url)
            self.referer = url
            homecontent = requests.get(url, headers=self.headers, timeout=self.timeout)
            tree = etree.HTML(homecontent.content)
            urlall = tree.xpath('//div[@class="pic"]/a/@href')
            for j in range(0,14):
                while urllist.full() is True:
                    pass
                try:
                    urllist.put(urlall[j])
                except:
                    print("页面内容一场，可能IP被暂时封")
    # 获取商家评论页内容
    def get_content(self):
        while True:
            # 获取商家评论页内容
            url=urllist.get()
            url = url+'/review_all'
            self.referer = url
            index_res = requests.get(url, headers=self.headers, timeout=self.timeout)
            content.put(index_res.text)
    # 解析html，输出css文件以及地址电话的svg链接
    def get_svg_html(self,content):
        # 正则匹配 css 文件
        result = re.search('<link rel="stylesheet" type="text/css" href="//s3plus(.*?)">', content, re.S)
        if result:
            css_url = 'http://s3plus' + result.group(1)
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.129 Safari/537.36'
            }
            css_res = requests.get(css_url, headers=headers)
            #print(f'css_url:{css_url}')
            css = css_res.text

            # 正则匹配商家地址使用的 svg 文件 url
            result = re.search('bb\[class.*?background-image: url\((.*?)\);', css, re.S)
            address_svg_url = 'http:' + result.group(1)
            address_svg = requests.get(address_svg_url, headers=headers).text
            #print(f'address_svg_url:{address_svg_url}')

            # 正则匹配商家电话号码使用的 svg 文件 url
            result = re.search('cc\[class.*?background-image: url\((.*?)\);', css, re.S)
            tell_svg_url = 'http:' + result.group(1)
            tell_svg = requests.get(tell_svg_url, headers=headers).text
            return css,address_svg,tell_svg
            #print(f'tell_svg_url:{tell_svg_url}')
    #标签里的字符从css中寻找出svg的坐标，再将原始html代码中的隐藏数据替换为svg中的正常字符
    def get_addr_font_map(self,content,css,address_svg):
        try:
            result = re.search('<bb class="(.*?)"></bb>', content, re.S)
            address_prefix = result.group(1)[:2]#这里（1代表正则匹配的第一个括号的结果（若在匹配多次的情况下，这里只有一个括号））[:2]截取字符串开头到第二个字符串
        except:
            print('地址类名不存在（页面未加密）\n')
            return content
        address_class_list = re.findall('\.%s(.*?){background:(.*?)px (.*?)px;}' % address_prefix, css, re.S)
        if 'textPath' in address_svg:#有两种不同写法的svg文件
            address_svg_y_list = re.findall('<path id="(\d+)" d="M0 (\d+) H600"/>', address_svg, re.S)
            address_result = re.findall('<textPath xlink:href="#(\d+)" textLength=".*?">(.*?)</textPath>', address_svg,re.S)
            address_words_dc = dict(address_result)
        else:
            address_result = re.findall('<text x=".*?" y="(.*?)">(.*?)</text>', address_svg, re.S)
            address_svg_y_list=[]
            for i in address_result:
                address_svg_y_list.append([i[0], i[0]])
            address_words_dc = dict(address_result)

        address_font_map = self.address_class_to_font(address_class_list, address_svg_y_list, address_words_dc,address_prefix)
        # 将 self.html 商铺地址加密的 class 样式替换成对应的中文字符
        address_class_set = re.findall('<bb class="(.*?)"></bb>', content, re.S)
        for class_name in address_class_set:
            content= re.sub('<bb class="{}"></bb>'.format(class_name), address_font_map[class_name], content)
        return content
    #同上，解析和替换电话字符
    def get_tell_font_map(self,content,css,tell_svg):
        try:
            result = re.search('<cc class="(.*?)"></cc>', content, re.S)
            #print(result.group(1))
            tell_prefix = result.group(1)[:2]
        except:
            print('电话类名不存在（页面未加密）\n')
            #print(self.html)
            return content
        tell_class_list = re.findall('\.%s(.*?){background:(.*?)px (.*?)px;}' % tell_prefix, css, re.S)
        tell_result = re.search('<text x="(.*?)" y=".*?">(.*?)</text>', tell_svg, re.S)
        tell_x_list = tell_result.group(1).split(' ')
        tell_words_str = tell_result.group(2)
        tell_font_map = self.tell_class_to_num(tell_class_list, tell_x_list, tell_words_str, tell_prefix)
        # 将 self.html 电话号码加密的 class 样式替换成对应的数字
        tell_class_set = re.findall('<cc class="(.*?)"></cc>', content, re.S)
        for class_name in tell_class_set:
            content = re.sub('<cc class="{}"></cc>'.format(class_name), tell_font_map[class_name], content)
        return content
    #将以上几个方法打包，实现传入原始html输出替换后的html
    def get_font_map(self):
        while True:
            # <bb class="xxx.*?"></bb>              地址
            # <cc class="xxx.*?"></cc>              电话
            # xxx 每天都会发生变化，所以动态匹配对应的前缀
            text=content.get()#获取html原始内容
            if '抱歉！页面无法访问......' in text:#若大众点评页面报错则抛出错误，跳过该页面
                print('页面出错')
                continue
            (css,address_svg,tell_svg)=self.get_svg_html(text)
            text=self.get_addr_font_map(text,css,address_svg)
            text=self.get_tell_font_map(text,css,tell_svg)
            alteredcontent.put(text)
            """
            匹配 css 文件中格式为 '.' + self.prefix + (.*?){background:(.*?)px (.*?)px;} 的 css 样式
    
            匹配 svg 文件中格式为 <path id="(\d+)" d="M0 (\d+) H600"/> 的字段，其中 id 的值对应 xlink:href="#(\d+)" 的值，
            d="M0 (\d+) H600" 的值对应 background中 y轴的偏移量
    
            匹配 svg 文件中格式为 <textPath xlink:href="#(\d+)" textLength=".*?">(.*?)</textPath> 的字段，(.*?) 对应一串中文字符串，
            最终的字符 = 中文字符串[x轴偏移量 / 字体大小]
            :return:
            """
    #替换地址字符的核心算法
    def address_class_to_font(self, class_list, y_list, words_dc, prefix):
        tmp_dc = dict()
        # 核心算法，将 css 转换为对应的字符
        for i in class_list:
            x_id = None
            for j in y_list:
                if int(j[1]) >= abs(int(float(i[2]))):
                    x_id = j[0]
                    break
            index = abs(int(float(i[1]))) // self.font_size
            tmp = words_dc[x_id][int(index)]
            tmp_dc[prefix + i[0]] = tmp
        return tmp_dc
    #替换电话数字的核心算法
    def tell_class_to_num(self, class_list, x_list, words_str, prefix):
        tmp_dc = dict()
        for i in class_list:
            x_index = None
            for index, num in enumerate(x_list):
                if int(num) >= abs(int(float(i[1]))):
                    x_index = index
                    break
            tmp = words_str[x_index]
            tmp_dc[prefix + i[0]] = tmp
        return tmp_dc
    #筛选出电话
    def phone_check(self,phone):
        #检查电话是否是手机号，若不是则丢弃
        for i in phone:
            result = re.match(self.phone_rule,i)
            if (result):
                checked_tell=i
                return checked_tell
            else:
                pass
        return False
    #获取地址电话并输出
    def get_shop_info(self):
        while True:
            #处理替换完后的正常数据
            #etree分析数据，拆分为店名电话和地址
            html=alteredcontent.get()   #获取修改后的html内容
            tree = etree.HTML(html)
            #获取地址
            shop_address = tree.xpath('//div[@class="address-info"]/text()')[0].replace('&nbsp;','').replace('\n','').replace(' ', '').split(':')[1]
            #确认是否有电话
            try:
                shop_tell = tree.xpath('//div[@class="phone-info"]/text()')[0].replace('&nbsp;','').replace('\n','').replace(' ', '').split(':')[1].split()
            except:
                print("缺少电话\n")
                alteredcontent.task_done()
                return False
            shop_name=tree.xpath('//h1[@class="shop-name"]/text()')[0]#获取店名
            #检测电话是否是手机号
            shop_tell=self.phone_check(shop_tell)
            if shop_tell==False:
                alteredcontent.task_done()
                print('无可用电话\n')
                continue
            #输出，并写入excel
            print(f'店名：{shop_name}')
            print(f'地址：{shop_address}\n电话：{shop_tell}\n')
            #print(alteredcontent.qsize())
            # 写入xlsx，通过锁来控制写入io
            lock.acquire()#获取io锁
            self.wbw.cell(self.flag, 1, shop_name)  # 往sheet中的写入数据
            self.wbw.cell(self.flag, 3, shop_address)
            self.wbw.cell(self.flag, 2, shop_tell)
            self.flag += 1
            lock.release()#写完后释放锁
            alteredcontent.task_done()


    def run(self):
        flag='Y'#是否继续运行的标志位
        while flag is 'Y':
            self.getCookies()
            self.homeurl= input('请输入搜索主页连接：')#这里的链接指的是指在大众点评搜索页的url，如：http://www.dianping.com/search/keyword/17/0_%E8%8A%B1%E5%BA%97/r1754o3
            self.maxpage = input('要多少页数据（每页15家店）？：')
            #self.homeurl = 'http://www.dianping.com/search/keyword/3/0_%E8%8A%B1%E5%BA%97/r59'
            starttime = time.time()
            url_thread = Thread(target=self.get_url)
            url_thread.daemon = True
            url_thread.start()

            for index in range(10):
                content_thread = Thread(target=self.get_content)
                content_thread.daemon = True
                content_thread.start()

            for index in range(10):
                alter_thread = Thread(target=self.get_font_map)
                alter_thread.daemon = True
                alter_thread.start()

            for index in range(10):
                out_thread = Thread(target=self.get_shop_info)
                out_thread.daemon = True
                out_thread.start()
            sleep(3)#等待3秒到alteredcontent队列中存在任务
            alteredcontent.join()
            self.wb.save(".\\result.xlsx")  # 保存
            #计时
            endtime = time.time()
            dtime = endtime - starttime
            print("爬取总共用时：%.8s s" % dtime)  # 显示到微秒
            flag=input('是否继续爬取(Y/N)')
if __name__ == '__main__':
    #创建队列
    urllist = Queue(maxsize=10)#页面url队列
    content=Queue(maxsize=10)#页面内容队列
    alteredcontent=Queue(maxsize=10)#替换后页面内容
    lock=threading.Lock()#创建写入锁
    dz =DaZhongDianPing()
    dz.run()